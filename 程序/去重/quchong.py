from xlrd import open_workbook
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws['A1'] = "位移"
ws['B1'] = "加速度"
ws['C1'] = "激光"
ws['D1'] = "速度"
ws['E1'] = "时间"

workbook = open_workbook('50-6.xlsx')   #打开指定文件
table = workbook.sheets()[0]   #打开第一张表格
n = table.nrows  #获取表格行数

cap0 = table.col_values(6)
a=0
for i in range(1, n):
    if cap0[i] != a:
        y = table.row_values(i)
        y[5] = y[5]/31.8+2.4795
        ws.append(y)
        a = cap0[i]

wb.save("50-6-1.xlsx")
