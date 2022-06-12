import modbus_tk
import modbus_tk.defines as cst
import modbus_tk.modbus_tcp as modbus_tcp
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
from time import time
from time import sleep
from datetime import datetime

"'数据采集去重(位移)'"
"'只采集一组数据'"

def DATA():
  if x[0] != z[0]:
    y = []
    if x[0] > 32767:
        x0 = float('%.3f' %((x[0] - 65536) * 0.001))
    else:
        x0 = float('%.3f' %(x[0] * 0.001))
    y.append(x0)
    y.append(float('%.5f' %dt))  # '%.3f' %  结果保留三位小数

    ws.append(y)


if __name__ == "__main__":
    try:
        # 连接MODBUS TCP从机
        master = modbus_tcp.TcpMaster(host="192.168.1.30")  # 构造主机对象，这里为TcpMaster对象，构造参数ip地址
        master.set_timeout(5.0)  # 等待时间

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "位移"
        ws['B1'] = "时间"

        z=[0]
        start = time()
        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 1)
            end = time()
            dt = end - start
            x = list(data)

            if dt <= 60:
                DATA()
                z = x
                sleep(0.01)
            else:
                wb.save("21.3.30.xlsx")
                break

    except:

        wb.save("21.3.30.xlsx")
        print(str(dt)+"秒时数据读取出现错误！")