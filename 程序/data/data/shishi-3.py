import modbus_tk
import modbus_tk.defines as cst
import modbus_tk.modbus_tcp as modbus_tcp
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
from time import time
from time import sleep
from datetime import datetime

def DATA():
    y = []
    for i in range(3):
        if x[i] > 32767:
            x0 = float('%.3f' % ((x[i] - 65536) * 0.001))
        else:
            x0 = float('%.3f' % (x[i] * 0.001))
        y.append(x0)
    y.append(float('%.5f' % dt))  # '%.3f' %  结果保留三位小数

    ws.append(y)

if __name__ == "__main__":
    try:
        # 连接MODBUS TCP从机
        master = modbus_tcp.TcpMaster(host="192.168.1.30")  # 构造主机对象，这里为TcpMaster对象，构造参数ip地址
        master.set_timeout(5.0)  # 等待时间

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "加速度"
        ws['B1'] = "位移"
        ws['C1'] = "速度"
        ws['D1'] = "时间"

        start = time()
        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 3)
            end = time()
            dt = end - start
            x = list(data)

            if dt <= 6:
                DATA()
                sleep(0.001)
            else:
                wb.save("data1.xlsx")
                break

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "加速度"
        ws['B1'] = "位移"
        ws['C1'] = "速度"
        ws['D1'] = "时间"
        start = time()
        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 3)
            end = time()
            dt = end - start
            x = list(data)

            if dt <= 6:
                DATA()
                sleep(0.001)
            else:
                wb.save("data2.xlsx")
                break

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "加速度"
        ws['B1'] = "位移"
        ws['C1'] = "速度"
        ws['D1'] = "时间"
        start = time()
        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 3)
            end = time()
            dt = end - start
            x = list(data)

            if dt <= 6:
                DATA()
                sleep(0.001)
            else:
                wb.save("data3.xlsx")
                break

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "加速度"
        ws['B1'] = "位移"
        ws['C1'] = "速度"
        ws['D1'] = "时间"
        start = time()
        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 3)
            end = time()
            dt = end - start
            x = list(data)

            if dt <= 6:
                DATA()
                sleep(0.001)
            else:
                wb.save("data4.xlsx")
                break

    except:
        wb.save("data1.xlsx")
        wb.save("data2.xlsx")
        wb.save("data3.xlsx")
        wb.save("data4.xlsx")





