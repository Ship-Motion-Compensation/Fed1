import modbus_tk
import modbus_tk.defines as cst
import modbus_tk.modbus_tcp as modbus_tcp
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
from time import time
from time import sleep
from datetime import datetime

def DATA():

        if x[0] > 32767:
            x0 = float('%.3f' % ((x[0] - 65536) * 0.001))
        else:
            x0 = float('%.3f' % (x[0] * 0.001))
        if x[1] > 32767:
            x1 = float('%.3f' % ((x[1] - 65536) * 0.001))
        else:
            x1 = float('%.3f' % (x[1] * 0.001))
        if x[2] > 32767:
            x2 = float('%.3f' % ((x[2] - 65536) * 0.001))
        else:
            x2 = float('%.3f' % (x[2] * 0.001))
        y = [x0, x1, x2, float('%.5f' % dt)]  # '%.3f' %  结果保留三位小数
        print(y)

        ws.append(y)

        # 添加数据

        y0_data.append(y[0])
        y1_data.append(y[1])
        y2_data.append(y[2])
        y3_data.append(y[3])


if __name__ == "__main__":
        # 连接MODBUS TCP从机
        master = modbus_tcp.TcpMaster(host="192.168.1.30")  #构造主机对象，这里为TcpMaster对象，构造参数ip地址
        master.set_timeout(5.0)  #等待时间

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "加速度"
        ws['B1'] = "位移"
        ws['C1'] = "速度"
        ws['D1'] = "时间"

        y0_data = []
        y1_data = []
        y2_data = []
        y3_data = []

        start = time()

        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 3)
            end = time()
            dt = end - start
            x = list(data)

            if dt <= 20:
                DATA()
                sleep(0.001)
            else:
                wb.save("data1.xlsx")
                # 画图
                plt.clf()  # 清除画布上的内容
                plt.rcParams['font.sans-serif'] = ['SimHei']  # 显示中文标签，不加的话，中文没法显示
                plt.rcParams['axes.unicode_minus'] = False  # 用来显示负号
                plt.subplot(3, 1, 1)  # 创建子图
                plt.xlabel('时间')
                plt.plot(y3_data, y0_data, 'r-', linewidth=1)
                plt.title('加速度')

                plt.subplot(3, 1, 2)
                plt.xlabel('时间')
                plt.plot(y3_data, y1_data, 'g-', linewidth=1)
                plt.title('位移')

                plt.subplot(3, 1, 3)
                plt.xlabel('时间')
                plt.plot(y3_data, y2_data, 'b-', linewidth=1)
                plt.title('位移')

                plt.show()
                break


