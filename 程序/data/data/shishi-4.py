import modbus_tk
import modbus_tk.defines as cst
import modbus_tk.modbus_tcp as modbus_tcp
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
from time import time
from time import sleep
from datetime import datetime

'"采集数据"'

f = "jg1.xlsx"  #数据保存到文件
n = 1000    #采集时间
def DATA():
    y = []
    for i in range(8):
        if x[i] > 32767:
            x0 = float('%.3f' %((x[i] - 65536) * 0.001))
        else:
            x0 = float('%.3f' %(x[i] * 0.001))
        y.append(x0)
    y.append(float('%.5f' %dt))  # '%.3f' %  结果保留三位小数

    ws.append(y)

if __name__ == "__main__":
    try:
        # 连接MODBUS TCP从机
        master = modbus_tcp.TcpMaster(host="192.168.1.30")  # 构造主机对象，这里为TcpMaster对象，构造参数ip地址
        master.set_timeout(5.0)  # 等待时间

        wb = Workbook()
        ws = wb.active
        ws['A1'] = "上平台中心位移"
        ws['B1'] = "上平台前侧位移"
        ws['C1'] = "横摇"
        ws['D1'] = "纵摇"
        ws['E1'] = "上平台角度p"
        ws['F1'] = "上平台角度r"
        ws['G1'] = "横摇激光"
        ws['H1'] = "纵摇激光"
        ws['I1'] = "时间"


        start = time()
        while True:
            # 读输入寄存器
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 64, 8)
            end = time()
            dt = end - start
            x = list(data)
            if dt <= n:
                DATA()
                sleep(0.01)
            else:
                wb.save(f)
                break

    except:

        wb.save(f)
        print(str(dt)+"秒时数据读取出现错误！")