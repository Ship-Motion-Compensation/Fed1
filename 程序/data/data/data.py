import modbus_tk
import modbus_tk.defines as cst
import modbus_tk.modbus_tcp as modbus_tcp
from openpyxl import load_workbook
from openpyxl import Workbook


#logger = modbus_tk.utils.create_logger("console")
if __name__ == "__main__":
     #try:
        # 连接MODBUS TCP从机
        master = modbus_tcp.TcpMaster(host="192.168.1.30")  #构造主机对象，这里为TcpMaster对象，构造参数ip地址
        master.set_timeout(5.0)  #等待时间
        #logger.info("connected")   #日志，info是级别，通常只记录关键节点信息，用于确认一切都是按照我们预期的那样进行工作
        # 读保持寄存器
        wb = Workbook()
        ws = wb.active
        i = 1
        x = []
        while(i<=500):
            i += 1
            #logger.info(master.execute(1, cst.READ_INPUT_REGISTERS, 64, 2))
            data = master.execute(1, cst.READ_INPUT_REGISTERS, 65, 3)
            x = list(data)
            x0 = '%.4f' % ((x[0] - 32768) / 32767 * 10)
            x1 = '%.4f' % ((x[1] - 32768) / 32767 * 10)
            x2 = '%.4f' % ((x[2] - 32768) / 32767 * 10)
            y = [x0,x1,x2]   #'%.3f' %  结果保留三位小数
            print(y)
            ws['A1'] = "加速度"
            ws['B1'] = "位移"
            ws['C1'] = "速度"
            ws.append(y)
        wb.save("data.xlsx")
        print(x)

     #except modbus_tk.modbus.ModbusError as e:
      #  logger.error("%s- Code=%d" % (e, e.get_exception_code()))

